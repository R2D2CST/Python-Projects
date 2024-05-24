# Importar librerías necesarias
# Importar la librería sys para gestionar el sistema
import sys
# Importar la librería os para interactuar con el sistema operativo
import os

# Importar librerías externas
# Importar la clase DocxTemplate de la librería docxtpl para trabajar con plantillas de Word
from docxtpl import DocxTemplate
# Importar la función load_workbook de la librería openpyxl para cargar archivos Excel
from openpyxl import load_workbook

# Cambiar al directorio del script
# Cambiar el directorio de trabajo al directorio del script actual
os.chdir(sys.path[0])

# Procedimiento para renderizar una plantilla
def render_template(template_path: str, output_path: str, context: dict) -> None:
    """
    Renderiza una plantilla de documento Word con el contexto dado y la guarda en el directorio de salida.
    
    Args:
        template_path (str): Ruta de la plantilla.
        output_path (str): Ruta de salida del documento renderizado.
        context (dict): Contexto con los datos para renderizar la plantilla.
    """
    # Cargar la plantilla
    document = DocxTemplate(template_path)
    # Renderizar el documento con el contexto dado
    document.render(context)
    # Guardar el documento renderizado en la ruta especificada
    document.save(output_path)

# Procedimiento para verificar la existencia de una ruta
def verify_path(path: str) -> None:
    """
    Verifica si la ruta especificada existe. Si no existe, crea la ruta.
    
    Args:
        path (str): Ruta a verificar.
    """
    # Verificar si la ruta no existe
    if not os.path.exists(path):
        # Crear la ruta
        os.makedirs(path)

# Procedimiento para obtener una lista de rutas de plantillas
def template_list() -> list[str]:
    """
    Obtiene una lista de rutas de plantillas en el directorio "Templates".
    
    Returns:
        list[str]: Lista de rutas de plantillas.
    """
    # Obtener las rutas de todas las plantillas en el directorio "Templates"
    return [os.path.join("Templates", file) for file in os.listdir("Templates")]

# Crear el directorio para los documentos renderizados
verify_path(path="rendered_templates")

# Abrir el archivo de Excel y obtener la hoja de cálculo
workbook = load_workbook("data_base.xlsx")
panel_ws = workbook["panel"]

# Extraer las llaves de la fila 2 (encabezados de las columnas)
keys = [cell.value for cell in panel_ws[2]]

# Iterar sobre cada fila de la hoja de cálculo y generar informes
for row in panel_ws.iter_rows(min_row=3, values_only=True):
    # Crear el diccionario de contexto para esta fila
    context = {keys[i]: value for i, value in enumerate(row)}
    
    # Crear la carpeta para el equipo si no existe
    verify_path(os.path.join("rendered_templates", str(context[keys[0]])))
    
    # Iterar sobre cada plantilla y renderizarla
    for template_path in template_list():
        # Construir la ruta de salida del documento renderizado
        output_path = os.path.join("rendered_templates", str(context[keys[0]]), f"rendered_{os.path.basename(template_path)}")
        # Renderizar la plantilla y guardar el documento renderizado
        render_template(template_path, output_path, context)
# Ultima Linea de código.