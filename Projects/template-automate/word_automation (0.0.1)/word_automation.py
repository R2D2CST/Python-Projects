# Native Python libraries
"""
import sys : used to change directory into the space where the python program is.
import os : used to change directory into the space where the python program is.
"""
import sys
import os

# Third party Python Libraries
"""
from openpyxl import load_workbook : used to load and create the Excel workbook object.
from docxtpl import DocxTemplate : used to create the word object template document.
from docxtpl import InlineImage : used to create the image placement objects in the word document.
from docx.shared import Cm, Inches, Mm, Emu : used for the unit handle.
import matplotlib.pyplot as plt : used to create the graph images to plot into the word document.
"""
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from docxtpl import InlineImage
from docx.shared import Cm
import matplotlib.pyplot as plt

# We change directory into the current path.
os.chdir (sys.path [0])

# Obtener el directorio actual del archivo Python
current_directory = os.getcwd()
# Nombre de la carpeta a verificar/crear
folder_name = "Graphs"
# Ruta completa de la carpeta
folder_path = os.path.join(current_directory, folder_name)
# Verificar si la carpeta existe
if not os.path.exists(folder_path):
    os.makedirs(folder_path) # Si no existe, crear la carpeta

# Load existing spreadsheet
wb = load_workbook ("word_automation.xlsm")
# We create an active worksheet
ws = wb ["Panel"]
# We select column A from active worksheet
Column_A = ws ["A"]
Column_B = ws ["B"]
# We get the values associated to each cell individually
Column_A_Values = []
for cell in Column_A:
    value = cell.value
    Column_A_Values.append (value)
Column_B_Values = []
for cell in Column_B:
    value = cell.value
    Column_B_Values.append (value)
# We transform individual values into a dictionary.
context = {}
length = len (Column_A_Values)
for index in range (0, length):
    context [Column_A_Values [index]] = Column_B_Values [index]

"""
We create the functions to generate the graphs according to the of graph we want to generate.
"""
def plot_lineGraph ()->None:
    """
    Procedure generates a line graph from the Line sheet with in the excel document
    """
    # Cargar el archivo de Excel
    wb = load_workbook("word_automation.xlsm")
    # Seleccionar la hoja "Graph"
    ws = wb["Line"]
    # Obtener los valores de las columnas A y B, excluyendo la primera fila y valores nulos.
    column_a = [cell.value for cell in ws["A"][1:]if cell.value is not None]
    column_b = [cell.value for cell in ws["B"][1:]if cell.value is not None]
    # Obtener los encabezados de las columnas A y B
    label_a = ws["A1"].value
    label_b = ws["B1"].value
    plt.figure(figsize=(8, 8))  # Definir el tamaño del gráfico
    # Crear el gráfico
    plt.plot(column_a, column_b)
    # Añadir etiquetas y título
    plt.xlabel(label_a)
    plt.ylabel(label_b)
    plt.title("Gráfico de " + label_a + " vs " + label_b)
    # Guardar el gráfico en una imagen PNG de 5 x 5 cm
    plt.savefig("Graphs/line_graph.png", dpi=300, format='png', bbox_inches='tight', pad_inches=0.1)
    pass

def plot_barGraph ()->None:
    """
    Procedure generates a bar graph from the Bar sheet with in the excel document
    """
    # Cargar el archivo de Excel
    wb = load_workbook("word_automation.xlsm")   
    # Seleccionar la hoja "Bars"
    ws = wb["Bars"]
    # Obtener los valores de las columnas A y B, excluyendo la primera fila y valores nulos.
    column_a = [cell.value for cell in ws["A"][1:]if cell.value is not None]
    column_b = [cell.value for cell in ws["B"][1:]if cell.value is not None]
    # Obtener las etiquetas de los ejes x e y
    xlabel = ws["A1"].value
    ylabel = ws["B1"].value
    plt.figure(figsize=(8, 8))  # Definir el tamaño del gráfico
    # Crear un gráfico de barras
    plt.bar(column_a, column_b)
    # Añadir etiquetas y título
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title("Graph/Gráfico de Barras")    
    # Guardar el gráfico de barras en una imagen PNG
    plt.savefig("Graphs/bar_graph.png", dpi=300, format='png', bbox_inches='tight', pad_inches=0.1)
    pass

def plot_pieGraph()->None:
    """
    Procedure generates a Pie graph from the Pie sheet with in the excel document
    """
    # Cargar el archivo de Excel
    wb = load_workbook("word_automation.xlsm")   
    # Seleccionar la hoja "Pie"
    ws = wb["Pie"]
    # Obtener los valores de las columnas A y B, excluyendo la primera fila y valores nulos.
    categories = [cell.value for cell in ws["A"][1:] if cell.value is not None]
    frequencies = [cell.value for cell in ws["B"][1:] if cell.value is not None]
    # Obtener las etiquetas de los ejes x e y
    xlabel = ws["A1"].value
    ylabel = ws["B1"].value
    # Crear un gráfico de pastel
    plt.figure(figsize=(8, 8))  # Definir el tamaño del gráfico
    plt.pie(frequencies, labels=categories, autopct='%1.1f%%')  # Crear el gráfico de pastel con porcentajes
    plt.title("Gráfico de Pastel")  # Agregar título
    plt.xlabel(xlabel)  # Agregar etiqueta del eje x
    plt.ylabel(ylabel)  # Agregar etiqueta del eje y
    # Guardar el gráfico de pastel en una imagen PNG
    plt.savefig("Graphs/pie_graph.png", dpi=300, format='png', bbox_inches='tight', pad_inches=0.1)
    pass

# We create the word template object.
document = DocxTemplate ("Template.docx")

# We create the graph to place into the file template.
plot_lineGraph ()
plot_barGraph ()
plot_pieGraph ()
# We create the image to place into the document.
placeholder_1 = InlineImage (document, "Graphs/line_graph.png", Cm (5))
placeholder_2 = InlineImage (document, "Graphs/bar_graph.png", Cm (5))
placeholder_3 = InlineImage (document, "Graphs/pie_graph.png", Cm (5))
# We append the image into the context dictionary.
context ["placeholder_1"] = placeholder_1
context ["placeholder_2"] = placeholder_2
context ["placeholder_3"] = placeholder_3

# We render the document template with our context values.
document.render (context)
# We save the rendered document into the same folder.
document.save ("Template_Rendered.docx")

# Last Line code